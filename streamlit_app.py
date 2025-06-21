import streamlit as st
import pandas as pd
import json
from io import BytesIO
from openai import OpenAI
import openpyxl

# API 키 확인

if “OPENAI_API_KEY” not in st.secrets:
st.error(“🚨 OpenAI API 키가 설정되지 않았습니다!”)
st.stop()

# ✅ OpenAI 클라이언트 초기화

client = OpenAI(api_key=st.secrets[“OPENAI_API_KEY”])

# 📂 엑셀 파일에서 모든 텍스트 추출 (개선된 버전)

def flatten_excel(file):
try:
# 파일 내용을 바이트스트림으로 읽기
file_content = file.getvalue()

```
    # openpyxl로 직접 읽기 (더 안정적)
    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    all_text = ''
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        st.info(f"📄 '{sheet_name}' 시트 처리 중...")
        
        for row in sheet.iter_rows(values_only=True):
            # None이 아닌 값들만 필터링하고 문자열로 변환
            row_text = []
            for cell in row:
                if cell is not None:
                    row_text.append(str(cell))
            
            if row_text:  # 빈 행이 아닐 때만 추가
                all_text += ' '.join(row_text) + '\n'
    
    workbook.close()
    
    if not all_text.strip():
        st.warning("⚠️ 엑셀 파일에서 텍스트를 찾을 수 없습니다.")
        return ""
    
    st.success(f"✅ 총 {len(all_text)} 글자 추출 완료!")
    return all_text.strip()
    
except Exception as e:
    st.error(f"❌ 엑셀 파일 읽기 오류: {str(e)}")
    
    # 대안으로 pandas 시도
    try:
        st.info("🔄 pandas로 재시도 중...")
        file_content = file.getvalue()
        xl = pd.ExcelFile(BytesIO(file_content))
        all_text = ''
        
        for sheet in xl.sheet_names:
            df = xl.parse(sheet, header=None)
            for _, row in df.iterrows():
                line = ' '.join([str(cell) for cell in row if pd.notna(cell)])
                if line.strip():
                    all_text += line + '\n'
        
        return all_text.strip() if all_text.strip() else ""
        
    except Exception as e2:
        st.error(f"❌ pandas로도 읽기 실패: {str(e2)}")
        return ""
```

# 🧠 GPT로 실무형 정보 추출

def extract_info_with_gpt(text):
try:
# 텍스트가 너무 길면 앞부분만 사용 (GPT 토큰 제한)
if len(text) > 10000:
text = text[:10000] + “\n[텍스트가 길어서 앞부분만 분석합니다]”
st.warning(“⚠️ 텍스트가 너무 길어 앞부분만 분석합니다.”)

```
    prompt = f"""
```

너는 여행사에서 사용하는 ‘골프여행 상품 견적 자동화 시스템’에 탑재된 GPT 분석기야.  
업로드된 상품 설명 문서에서 **아래 JSON 구조**에 맞춰 필요한 정보를 정확하게 추출해줘.

반드시 지켜야 할 규칙은 다음과 같아

1. 모든 항목을 JSON 포맷으로 출력해. 텍스트나 설명 말고 JSON만 반환해.
1. 항목이 없는 경우에도 null 또는 빈 배열 []로 명시적으로 채워.
1. 숫자는 따옴표 없이 숫자형(int)으로 출력해.
1. “price_by_date”는 날짜별 가격이 있다면 최대 5개까지 정리해주고, 없다면 빈 배열로.
1. “flight_included”는 항공이 포함되었는지 문맥을 이해해서 true/false로 정확하게 판단해.
1. 포함/불포함 항목은 문장이나 표로 표현된 내용까지 감지해서 배열로 정리해.

출력 JSON 형식은 아래와 같아. 이 형식을 그대로 유지해줘:

{{
“product_name”: “”,
“region”: “”,
“departure_date”: “”,
“duration”: “”,
“rounds_per_day”: “”,
“total_rounds”: “”,
“price”: null,
“price_by_date”: [
{{
“date”: “2025-07-01”,
“price”: 1390000
}},
{{
“date”: “2025-07-08”,
“price”: 1490000
}}
],
“flight_included”: false,
“includes”: [],
“excludes”: [],
“hotel”: “”,
“golf_courses”: [],
“transport”: “”,
“extra_charges”: {{
“single_charge”: null,
“caddie_fee”: null,
“cart_fee”: null
}}
}}

이제 아래 텍스트를 분석해서 위 JSON 형식으로 추출해줘:

{text}
“””
response = client.chat.completions.create(
model=“gpt-4”,
messages=[{“role”: “user”, “content”: prompt}],
temperature=0,
max_tokens=2000
)
return response.choices[0].message.content

```
except Exception as e:
    st.error(f"❌ GPT API 호출 오류: {str(e)}")
    return ""
```

# 🖥️ Streamlit 웹앱 UI 구성

st.set_page_config(page_title=“⛳ 골프상품 자동추출기”, layout=“wide”)
st.title(“⛳ 실무용 골프 여행 상품 자동 추출기”)

# 파일 업로드 안내

st.info(“📋 **사용법**: 엑셀 파일(.xlsx, .xls)을 업로드하면 자동으로 골프 상품 정보를 추출합니다.”)

uploaded = st.file_uploader(“📂 엑셀 파일을 업로드하세요”, type=[“xls”, “xlsx”])

if uploaded:
# 파일 정보 표시
st.success(f”📁 파일 업로드 완료: {uploaded.name} ({uploaded.size:,} bytes)”)

```
with st.spinner("📖 엑셀 파일을 읽는 중..."):
    text = flatten_excel(uploaded)

if text:  # 텍스트가 있을 때만 표시
    with st.expander("📋 추출된 텍스트 미리보기", expanded=False):
        preview_text = text[:1000] + "..." if len(text) > 1000 else text
        st.text_area("", preview_text, height=200)
    
    if st.button("🧠 GPT로 정보 추출", type="primary"):
        with st.spinner("🤖 GPT가 내용을 분석 중입니다..."):
            result = extract_info_with_gpt(text)
        
        if result:  # 결과가 있을 때만 처리
            st.subheader("📦 GPT 응답")
            
            # JSON 코드 블록에서 JSON만 추출
            if "```json" in result:
                json_start = result.find("```json") + 7
                json_end = result.find("```", json_start)
                if json_end != -1:
                    result = result[json_start:json_end].strip()
            
            st.code(result, language="json")
            
            try:
                parsed = json.loads(result)
                st.success("✅ JSON 파싱 성공!")
                
                # 예쁘게 표시
                st.subheader("📊 추출된 정보")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**기본 정보**")
                    st.write(f"🏷️ **상품명**: {parsed.get('product_name', 'N/A')}")
                    st.write(f"🌍 **지역**: {parsed.get('region', 'N/A')}")
                    st.write(f"📅 **출발일**: {parsed.get('departure_date', 'N/A')}")
                    st.write(f"⏰ **기간**: {parsed.get('duration', 'N/A')}")
                    st.write(f"⛳ **총 라운드**: {parsed.get('total_rounds', 'N/A')}")
                
                with col2:
                    st.write("**가격 정보**")
                    if parsed.get('price'):
                        st.write(f"💰 **가격**: {parsed['price']:,}원")
                    
                    if parsed.get('price_by_date'):
                        st.write("📅 **날짜별 가격**:")
                        for item in parsed['price_by_date']:
                            st.write(f"  - {item.get('date')}: {item.get('price', 0):,}원")
                
                # 전체 JSON 표시
                with st.expander("📄 전체 JSON 데이터", expanded=False):
                    st.json(parsed)
                    
            except json.JSONDecodeError as e:
                st.error(f"❌ JSON 파싱 실패: {str(e)}")
                st.write("**원본 응답:**")
                st.text(result)
else:
    st.warning("⚠️ 엑셀 파일에서 텍스트를 추출할 수 없습니다. 파일 형식을 확인해주세요.")
```